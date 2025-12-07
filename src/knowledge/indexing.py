import asyncio
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import aiofiles
from langchain_community.document_loaders import (
    CSVLoader,
    JSONLoader,
    PyPDFLoader,
    TextLoader,
    UnstructuredHTMLLoader,
    UnstructuredMarkdownLoader,
    UnstructuredWordDocumentLoader,
)
from langchain_text_splitters import RecursiveCharacterTextSplitter

from src.knowledge.utils import calculate_content_hash
from src.storage.minio import get_minio_client
from src.utils import hashstr, logger

SUPPORTED_FILE_EXTENSIONS: tuple[str, ...] = (
    ".txt",
    ".md",
    ".doc",
    ".docx",
    ".html",
    ".htm",
    ".json",
    ".csv",
    ".xls",
    ".xlsx",
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".bmp",
    ".tiff",
    ".tif",
    ".zip",
)


def is_supported_file_extension(file_name: str | os.PathLike[str]) -> bool:
    """Check whether the given file path has a supported extension."""
    return Path(file_name).suffix.lower() in SUPPORTED_FILE_EXTENSIONS


def _make_unique_columns(columns: list) -> list:
    """
    处理重复的列名,给重复的列添加数字后缀

    Args:
        columns: 原始列名列表

    Returns:
        处理后的唯一列名列表

    Example:
        ['A', 'B', 'A', 'C', 'B'] -> ['A', 'B', 'A_2', 'C', 'B_2']
    """
    if not columns:
        return columns

    seen = {}
    unique_columns = []

    for col in columns:
        # 处理 None 或空值
        if col is None or (isinstance(col, str) and not col.strip()):
            col = "Unnamed"

        # 转换为字符串
        col_str = str(col)

        # 如果这个列名已经出现过
        if col_str in seen:
            seen[col_str] += 1
            # 添加后缀
            unique_col = f"{col_str}_{seen[col_str]}"
            unique_columns.append(unique_col)
        else:
            # 第一次出现
            seen[col_str] = 1
            unique_columns.append(col_str)

    return unique_columns


def _extract_word_text(file_path: Path) -> str:
    """
    Parse Word documents (.doc/.docx) into plain text.

    Try python-docx first for docx files and fall back to the unstructured
    loader so legacy .doc files are still parsed when possible.
    """
    try:
        from docx import Document  # type: ignore

        doc = Document(file_path)
        text = "\n".join(paragraph.text for paragraph in doc.paragraphs).strip()
        if text:
            return text
    except Exception as docx_error:  # noqa: BLE001
        logger.warning(f"python-docx failed to parse {file_path.name}: {docx_error}")

    try:
        loader = UnstructuredWordDocumentLoader(str(file_path))
        docs = loader.load()
        return "\n".join(doc.page_content for doc in docs).strip()
    except Exception as unstructured_error:  # noqa: BLE001
        logger.error(f"Unstructured failed to parse {file_path.name}: {unstructured_error}")
        raise ValueError(f"无法解析 Word 文档: {file_path.name}") from unstructured_error


def _extract_docx_markdown_with_images(file_path: Path, params: dict | None = None) -> str:
    params = params or {}
    db_id = params.get("db_id") or "word-docs"
    minio_client = get_minio_client()
    bucket_name = "kb-images"
    minio_client.ensure_bucket_exists(bucket_name)
    file_id = hashstr(file_path.name, length=16)

    with zipfile.ZipFile(file_path, "r") as zf:
        rels_path = "word/_rels/document.xml.rels"
        rid_to_target: dict[str, str] = {}
        try:
            rels_xml = zf.read(rels_path).decode("utf-8")
            rels_root = ET.fromstring(rels_xml)
            for rel in list(rels_root):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                rtype = rel.attrib.get("Type")
                if rid and target and rtype and rtype.endswith("/image"):
                    rid_to_target[rid] = target
        except Exception:
            rid_to_target = {}

        doc_xml = zf.read("word/document.xml").decode("utf-8")
        ns = {
            "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        root = ET.fromstring(doc_xml)
        md_lines: list[str] = []

        for p in root.findall(".//w:p", ns):
            texts = [t.text or "" for t in p.findall(".//w:t", ns)]
            para_text = "".join(texts).strip()
            image_urls: list[tuple[str, str]] = []

            for blip in p.findall(".//a:blip", ns):
                rid = blip.attrib.get(f"{{{ns['r']}}}embed")
                if not rid:
                    continue
                target = rid_to_target.get(rid)
                if not target:
                    continue
                media_path = target if target.startswith("word/") else f"word/{target}"
                try:
                    data = zf.read(media_path)
                    object_name = f"{db_id}/{file_id}/images/{Path(target).name}"
                    suffix = Path(target).suffix.lower()
                    content_type = {
                        ".jpg": "image/jpeg",
                        ".jpeg": "image/jpeg",
                        ".png": "image/png",
                        ".gif": "image/gif",
                        ".webp": "image/webp",
                        ".bmp": "image/bmp",
                        ".tif": "image/tiff",
                        ".tiff": "image/tiff",
                    }.get(suffix, "image/jpeg")
                    result = minio_client.upload_file(
                        bucket_name=bucket_name,
                        object_name=object_name,
                        data=data,
                        content_type=content_type,
                    )
                    image_urls.append((Path(target).name, result.url))
                except Exception as e:
                    logger.error(f"上传图片失败 {Path(target).name}: {e}")
                    continue

            line = para_text
            for name, url in image_urls:
                line = f"{line}\n![{name}]({url})"
            if line:
                md_lines.append(line)

    return "\n\n".join(md_lines)


def chunk_with_parser(file_path, params=None):
    """
    使用文件解析器将文件切分成固定大小的块

    Args:
        file_path: 文件路径
        params: 参数
    """
    params = params or {}
    chunk_size = int(params.get("chunk_size", 500))
    chunk_overlap = int(params.get("chunk_overlap", 100))

    file_type = Path(file_path).suffix.lower()

    # 选择合适的加载器
    if file_type in [".txt"]:
        loader = TextLoader(file_path)

    elif file_type in [".md"]:
        loader = UnstructuredMarkdownLoader(file_path)

    elif file_type in [".docx", ".doc"]:
        loader = UnstructuredWordDocumentLoader(file_path)

    elif file_type in [".html", ".htm"]:
        loader = UnstructuredHTMLLoader(file_path)

    elif file_type in [".json"]:
        loader = JSONLoader(file_path, jq_schema=".")

    elif file_type in [".csv"]:
        loader = CSVLoader(file_path)

    else:
        raise ValueError(f"不支持的文件类型: {file_type}")

    # 加载文档
    docs = loader.load()

    # 创建文本分割器
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ".", " ", ""],
    )

    # 分割文档
    nodes = text_splitter.split_documents(docs)

    # 添加序号信息到metadata
    for i, node in enumerate(nodes):
        if node.metadata is None:
            node.metadata = {}
        node.metadata["chunk_idx"] = i

    return nodes


def chunk_text(text, params=None):
    """
    将文本切分成固定大小的块
    """
    params = params or {}
    chunk_size = int(params.get("chunk_size", 500))
    chunk_overlap = int(params.get("chunk_overlap", 100))

    # 创建文本分割器
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size, chunk_overlap=chunk_overlap, separators=["\n\n", "\n", ".", " ", ""]
    )

    # 分割文档
    nodes = text_splitter.split_text(text)

    # 添加序号信息到metadata
    nodes = [{"text": node, "metadata": {"chunk_idx": i}} for i, node in enumerate(nodes)]
    return nodes


def chunk(text_or_path, params=None):
    raise NotImplementedError("chunk is deprecated, use chunk_with_parser or chunk_text instead")


def pdfreader(file_path, params=None):
    """读取PDF文件并返回text文本"""
    if isinstance(file_path, str):
        file_path = Path(file_path)

    assert file_path.exists(), "File not found"
    assert file_path.suffix.lower() == ".pdf", "File format not supported"

    # 使用LangChain的PDF加载器
    loader = PyPDFLoader(str(file_path))
    docs = loader.load()

    # 简单的拼接起来之后返回纯文本
    text = "\n\n".join([d.page_content for d in docs])
    return text


def plainreader(file_path):
    """读取普通文本文件并返回text文本"""
    assert os.path.exists(file_path), "File not found"

    # 使用LangChain的文本加载器
    loader = TextLoader(str(file_path))
    docs = loader.load()
    text = "\n\n".join([d.page_content for d in docs])
    return text


def parse_pdf(file, params=None):
    """
    解析PDF文件，支持多种OCR方式

    Args:
        file: PDF文件路径
        params: 参数字典，包含enable_ocr设置

    Returns:
        str: 解析得到的文本

    Raises:
        DocumentProcessorException: 处理失败时抛出
    """
    from src.plugins.document_processor_base import DocumentProcessorException
    from src.plugins.document_processor_factory import DocumentProcessorFactory

    params = params or {}
    opt_ocr = params.get("enable_ocr", "disable")

    if opt_ocr == "disable":
        return pdfreader(file, params=params)

    try:
        return DocumentProcessorFactory.process_file(opt_ocr, file, params)

    except DocumentProcessorException as e:
        logger.error(f"文档处理失败: {e.service_name} - {str(e)}")
        raise
    except Exception as e:
        logger.error(f"PDF 解析失败: {str(e)}")
        raise DocumentProcessorException(f"PDF解析失败: {str(e)}", opt_ocr, "parsing_failed")


def parse_image(file, params=None):
    """
    解析图像文件，支持多种OCR方式

    Args:
        file: 图像文件路径
        params: 参数字典，包含enable_ocr设置

    Returns:
        str: 解析得到的文本

    Raises:
        DocumentProcessorException: 处理失败时抛出
        ValueError: 图像文件禁用OCR时抛出
    """
    from src.plugins.document_processor_base import DocumentProcessorException
    from src.plugins.document_processor_factory import DocumentProcessorFactory

    params = params or {}
    opt_ocr = params.get("enable_ocr", "disable")

    # 图像文件必须使用 OCR,不能禁用
    if opt_ocr == "disable":
        raise ValueError(
            "图像文件必须启用OCR才能提取文本内容。"
            "请选择OCR方式 (onnx_rapid_ocr/mineru_ocr/mineru_official/paddlex_ocr) 或移除该文件。"
        )

    try:
        return DocumentProcessorFactory.process_file(opt_ocr, file, params)

    except DocumentProcessorException as e:
        logger.error(f"图像处理失败: {e.service_name} - {str(e)}")
        raise
    except Exception as e:
        logger.error(f"图像解析失败: {str(e)}")
        raise DocumentProcessorException(f"图像解析失败: {str(e)}", opt_ocr, "parsing_failed")


async def parse_pdf_async(file, params=None):
    return await asyncio.to_thread(parse_pdf, file, params=params)


async def parse_image_async(file, params=None):
    return await asyncio.to_thread(parse_image, file, params=params)


async def process_file_to_markdown(file_path: str, params: dict | None = None) -> str:
    """
    将不同类型的文件转换为markdown格式 - 支持本地文件和MinIO文件

    Args:
        file_path: 文件路径或MinIO URL
        params: 处理参数，对于ZIP文件需要包含 db_id

    Returns:
        markdown格式内容

    Note:
        对于ZIP文件，会在params中保存处理结果供调用方使用：
        - params['_zip_images_info']: 图片信息列表
        - params['_zip_content_hash']: 内容哈希值
    """
    import tempfile
    import aiofiles
    import os

    # 检测是否是MinIO URL
    from src.knowledge.utils.kb_utils import is_minio_url

    if is_minio_url(file_path):
        # 从MinIO下载文件到临时位置
        logger.debug(f"Downloading file from MinIO: {file_path}")

        # 从MinIO URL中提取文件名
        if "?" in file_path:
            file_path_clean = file_path.split("?")[0]
        else:
            file_path_clean = file_path

        original_filename = file_path_clean.split("/")[-1]

        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(original_filename).suffix) as temp_file:
            temp_path = temp_file.name

        try:
            # 使用通用函数解析MinIO URL并下载文件
            from src.knowledge.utils.kb_utils import parse_minio_url
            from src.storage.minio.client import get_minio_client

            # 解析MinIO URL获取bucket_name和object_name
            bucket_name, object_name = parse_minio_url(file_path)

            # 获取MinIO客户端并下载文件
            minio_client = get_minio_client()
            file_content = await minio_client.adownload_file(bucket_name, object_name)

            # 写入临时文件
            async with aiofiles.open(temp_path, "wb") as f:
                await f.write(file_content)

            logger.debug(f"File downloaded to temp path: {temp_path}")

            # 使用临时文件路径
            actual_file_path = temp_path

        except Exception as e:
            # 清理临时文件
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            logger.error(f"Failed to download file from MinIO: {e}")
            raise ValueError(f"无法从MinIO下载文件: {e}")
    else:
        # 本地文件
        actual_file_path = file_path

    try:
        file_path_obj = Path(actual_file_path)
        file_ext = file_path_obj.suffix.lower()
        original_filename = file_path_obj.name

        if file_ext == ".pdf":
            # 使用 OCR 处理 PDF
            text = await parse_pdf_async(str(file_path_obj), params=params)
            result = f"{text}"

        elif file_ext in [".txt", ".md"]:
            # 直接读取文本文件
            with open(file_path_obj, encoding="utf-8") as f:
                content = f.read()
            result = f"{content}"

        elif file_ext == ".docx":
            text = _extract_docx_markdown_with_images(file_path_obj, params=params)
            result = f"" + text

        elif file_ext == ".doc":
            text = _extract_word_text(file_path_obj)
            result = f"{text}"

        elif file_ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"]:
            # 使用 OCR 处理图片
            text = await parse_image_async(str(file_path_obj), params=params)
            result = f"{text}"

        elif file_ext in [".html", ".htm"]:
            # 使用 BeautifulSoup 处理 HTML 文件
            from markdownify import markdownify as md

            with open(file_path_obj, encoding="utf-8") as f:
                content = f.read()
            text = md(content, heading_style="ATX")
            result = f"{text}"

        elif file_ext == ".csv":
            # 处理 CSV 文件
            import pandas as pd

            df = pd.read_csv(file_path_obj)
            # 将每一行数据与表头组合成独立的表格
            markdown_content = f""

            for index, row in df.iterrows():
                # 创建包含表头和当前行的小表格
                row_df = pd.DataFrame([row], columns=df.columns)
                markdown_table = row_df.to_markdown(index=False)
                markdown_content += f"{markdown_table}\n\n"

            result = markdown_content.strip()

        elif file_ext in [".xls", ".xlsx"]:
            # 处理 Excel 文件
            import pandas as pd
            from openpyxl import load_workbook

            markdown_content = f""

            # 使用 openpyxl 加载工作簿以正确处理合并单元格
            wb = load_workbook(file_path_obj, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 先取消所有合并单元格，并填充值
                merged_ranges = list(ws.merged_cells.ranges)

                for merged_range in merged_ranges:
                    # 获取合并区域左上角单元格的值
                    min_row, min_col, max_row, max_col = (
                        merged_range.min_row,
                        merged_range.min_col,
                        merged_range.max_row,
                        merged_range.max_col,
                    )

                    # 获取左上角单元格的值
                    top_left_value = ws.cell(row=min_row, column=min_col).value

                    # 取消合并
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

                    # 在所有原合并单元格区域填充值
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col).value = top_left_value

                # 转换为DataFrame
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)

                # 第一行作为列名
                columns = data[0] if data else []
                df_data = data[1:] if len(data) > 1 else []

                # 处理重复的列名,给重复的列添加后缀
                columns = _make_unique_columns(columns)

            df = pd.DataFrame(df_data, columns=columns)

            markdown_content += f"## {sheet_name}\n\n"

            # 在最左列添加标题字段
            table_title = f"{file_path_obj.stem} - {sheet_name}"  # 使用"文件名 - Sheet名"作为标题
            df.insert(0, "表格标题", table_title)

            # 将每10行数据与表头组合成独立的表格
            chunk_size = 10
            for i in range(0, len(df), chunk_size):
                # 获取当前10行数据(或剩余不足10行的数据)
                chunk_df = df.iloc[i : i + chunk_size]
                markdown_content += f"### 数据行 {i + 1}-{min(i + chunk_size, len(df))}\n\n"
                markdown_table = chunk_df.to_markdown(index=False)
                markdown_content += f"{markdown_table}\n\n"

            result = markdown_content.strip()

        elif file_ext == ".json":
            # 处理 JSON 文件
            import json

            async with aiofiles.open(file_path_obj, encoding="utf-8") as f:
                content = await f.read()
            data = json.loads(content)
            # 将 JSON 数据格式化为 markdown 代码块
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            result = f"```json\n{json_str}\n```"

        elif file_ext == ".zip":
            if not params or "db_id" not in params:
                raise ValueError("ZIP文件处理需要在params中提供db_id参数")

            zip_result = await asyncio.to_thread(_process_zip_file, str(file_path_obj), params["db_id"])

            # 将处理结果保存到params中供调用方使用
            params["_zip_images_info"] = zip_result["images_info"]
            params["_zip_content_hash"] = zip_result["content_hash"]

            result = zip_result["markdown_content"]

        else:
            # 尝试作为文本文件读取
            raise ValueError(f"Unsupported file type: {file_ext}")

    except Exception as e:
        # 清理临时文件
        if is_minio_url(file_path) and os.path.exists(actual_file_path):
            try:
                os.unlink(actual_file_path)
                logger.debug(f"Cleaned up temp file: {actual_file_path}")
            except Exception as cleanup_e:
                logger.warning(f"Failed to clean up temp file {actual_file_path}: {cleanup_e}")
        raise

    finally:
        # 清理临时文件
        if is_minio_url(file_path) and os.path.exists(actual_file_path):
            try:
                os.unlink(actual_file_path)
                logger.debug(f"Cleaned up temp file: {actual_file_path}")
            except Exception as e:
                logger.warning(f"Failed to clean up temp file {actual_file_path}: {e}")

    return result


def _process_zip_file(zip_path: str, db_id: str) -> dict:
    """
    处理ZIP文件，提取markdown内容和图片（内部函数）

    Args:
        zip_path: ZIP文件路径
        db_id: 数据库ID

    Returns:
        dict: {
            "markdown_content": str,      # markdown内容
            "content_hash": str,         # 内容哈希值
            "images_info": list[dict]    # 图片信息列表
        }

    Raises:
        FileNotFoundError: ZIP文件不存在
        ValueError: ZIP文件格式错误或内容不符合要求
    """
    # 1. 安全检查
    if not os.path.exists(zip_path):
        raise FileNotFoundError(f"ZIP 文件不存在: {zip_path}")

    # 2. 解压ZIP并提取内容
    with zipfile.ZipFile(zip_path, "r") as zf:
        # 安全检查：防止路径遍历攻击
        for name in zf.namelist():
            if name.startswith("/") or name.startswith("\\"):
                raise ValueError(f"ZIP 包含不安全路径: {name}")
            if ".." in Path(name).parts:
                raise ValueError(f"ZIP 路径包含上级引用: {name}")

        # 查找markdown文件
        md_files = [n for n in zf.namelist() if n.lower().endswith(".md")]
        if not md_files:
            raise ValueError("压缩包中未找到 .md 文件")

        # 优先使用 full.md，否则使用第一个md文件
        md_file = next((n for n in md_files if Path(n).name == "full.md"), md_files[0])

        # 读取markdown内容
        with zf.open(md_file) as f:
            markdown_content = f.read().decode("utf-8")

        # 3. 处理图片
        images_info = []
        images_dir = _find_images_directory(zf, md_file)

        if images_dir:
            images_info = _process_images(zf, images_dir, db_id, md_file)
            markdown_content = _replace_image_links(markdown_content, images_info)

    # 4. 生成结果
    content_hash = asyncio.run(calculate_content_hash(markdown_content.encode("utf-8")))

    return {
        "markdown_content": markdown_content,
        "content_hash": content_hash,
        "images_info": images_info,
    }


def _find_images_directory(zip_file: zipfile.ZipFile, md_file_path: str) -> str | None:
    """查找images目录"""
    md_parent = Path(md_file_path).parent

    # 候选目录
    candidates = []
    if str(md_parent) != ".":
        candidates.extend([str(md_parent / "images"), str(md_parent.parent / "images")])
    candidates.append("images")

    # 查找存在的目录
    for cand in candidates:
        cand_clean = cand.rstrip("/")
        if any(n.startswith(cand_clean + "/") for n in zip_file.namelist()):
            return cand_clean

    return None


async def _process_images(zip_file: zipfile.ZipFile, images_dir: str, db_id: str, md_file_path: str) -> list[dict]:
    """处理图片：上传到MinIO并返回信息"""
    # 支持的图片格式
    SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
    CONTENT_TYPE_MAP = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
    }

    images = []
    image_names = [n for n in zip_file.namelist() if n.startswith(images_dir + "/")]

    # 上传图片到MinIO
    minio_client = await get_minio_client()
    bucket_name = "kb-images"
    minio_client.ensure_bucket_exists(bucket_name)

    file_id = hashstr(Path(md_file_path).name, length=16)

    for img_name in image_names:
        suffix = Path(img_name).suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            continue

        try:
            # 读取图片数据
            with zip_file.open(img_name) as f:
                data = f.read()

            # 上传到MinIO
            object_name = f"{db_id}/{file_id}/images/{Path(img_name).name}"
            content_type = CONTENT_TYPE_MAP.get(suffix, "image/jpeg")

            result = minio_client.upload_file(
                bucket_name=bucket_name,
                object_name=object_name,
                data=data,
                content_type=content_type,
            )

            # 记录图片信息
            img_info = {"name": Path(img_name).name, "url": result.url, "path": f"images/{Path(img_name).name}"}
            images.append(img_info)

            logger.debug(f"图片上传成功: {Path(img_name).name} -> {result.url}")

        except Exception as e:
            logger.error(f"上传图片失败 {Path(img_name).name}: {e}")
            continue

    return images


def _replace_image_links(markdown_content: str, images: list[dict]) -> str:
    """替换markdown中的图片链接为MinIO URL"""
    if not images:
        return markdown_content

    # 构建路径映射
    image_map = {}
    for img in images:
        path = img["path"]
        url = img["url"]
        image_map[path] = url
        image_map[f"/{path}"] = url
        image_map[img["name"]] = url

    def replace_link(match):
        alt_text = match.group(1) or ""
        img_path = match.group(2)

        # 尝试匹配各种路径格式
        for pattern, url in image_map.items():
            if img_path.endswith(pattern) or img_path == pattern:
                return f"![{alt_text}]({url})"

        # 尝试文件名匹配
        filename = os.path.basename(img_path)
        if filename in image_map:
            return f"![{alt_text}]({image_map[filename]})"

        return match.group(0)

    # 使用正则表达式替换图片链接
    pattern = r"!\[([^\]]*)\]\(([^)]+)\)"
    return re.sub(pattern, replace_link, markdown_content)


async def process_url_to_markdown(url: str, params: dict | None = None) -> str:
    raise NotImplementedError("URL 解析功能已禁用")
