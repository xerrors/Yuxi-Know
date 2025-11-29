"""
测试文档索引处理的变更
文件: src/knowledge/indexing.py
"""

import os
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pytest
import pandas as pd
from openpyxl import Workbook
from src.knowledge.indexing import process_file_to_markdown, _make_unique_columns


class TestMakeUniqueColumns:
    """测试新增的 _make_unique_columns 函数"""

    def test_no_duplicates(self):
        """测试没有重复列名的情况"""
        columns = ["列A", "列B", "列C", "列D"]
        result = _make_unique_columns(columns)

        assert result == columns
        print(f"✓ 无重复列名处理正确: {result}")

    def test_with_duplicates(self):
        """测试有重复列名的情况"""
        columns = ["姓名", "年龄", "姓名", "地址", "年龄", "姓名"]
        result = _make_unique_columns(columns)

        expected = ["姓名", "年龄", "姓名_2", "地址", "年龄_2", "姓名_3"]
        assert result == expected
        print(f"✓ 重复列名处理正确: {result}")

    def test_with_none_values(self):
        """测试包含 None 值的列名"""
        columns = ["A", None, "B", None, "C"]
        result = _make_unique_columns(columns)

        assert result[0] == "A"
        assert result[1] == "Unnamed"
        assert result[2] == "B"
        assert result[3] == "Unnamed_2"
        assert result[4] == "C"
        print(f"✓ None 值处理正确: {result}")

    def test_with_empty_strings(self):
        """测试包含空字符串的列名"""
        columns = ["A", "", "B", "  ", "C"]
        result = _make_unique_columns(columns)

        # 空字符串应该被转换为 "Unnamed"
        assert result[0] == "A"
        assert result[1] == "Unnamed"
        assert result[2] == "B"
        assert result[3] == "Unnamed_2"
        assert result[4] == "C"
        print(f"✓ 空字符串处理正确: {result}")

    def test_empty_list(self):
        """测试空列表"""
        columns = []
        result = _make_unique_columns(columns)

        assert result == []
        print(f"✓ 空列表处理正确")

    def test_all_same_names(self):
        """测试所有列名相同的情况"""
        columns = ["数据", "数据", "数据", "数据"]
        result = _make_unique_columns(columns)

        expected = ["数据", "数据_2", "数据_3", "数据_4"]
        assert result == expected
        print(f"✓ 全部相同列名处理正确: {result}")


class TestExcelProcessingChanges:
    """测试 Excel 处理逻辑的变更"""

    @pytest.mark.asyncio
    async def test_excel_basic_processing(self, tmp_path):
        """测试基础 Excel 处理"""
        # 创建测试 Excel 文件
        test_file = tmp_path / "test_basic.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "测试表"

        # 添加表头
        ws["A1"] = "姓名"
        ws["B1"] = "年龄"
        ws["C1"] = "城市"

        # 添加数据
        data = [
            ["张三", 25, "北京"],
            ["李四", 30, "上海"],
            ["王五", 28, "广州"],
        ]
        for i, row in enumerate(data, start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果
        assert result is not None
        assert "测试表" in result
        assert "表格标题" in result  # 新增的标题列
        assert "姓名" in result
        assert "张三" in result
        print(f"✓ 基础 Excel 处理成功")
        print(f"处理结果预览:\n{result[:300]}...")

    @pytest.mark.asyncio
    async def test_excel_merged_cells(self, tmp_path):
        """测试包含合并单元格的 Excel 处理"""
        test_file = tmp_path / "test_merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "合并单元格测试"

        # 添加表头
        ws["A1"] = "部门"
        ws["B1"] = "员工"
        ws["C1"] = "工资"

        # 添加数据并合并单元格
        ws["A2"] = "技术部"
        ws.merge_cells("A2:A4")  # 合并 A2:A4

        ws["B2"] = "员工1"
        ws["B3"] = "员工2"
        ws["B4"] = "员工3"

        ws["C2"] = 10000
        ws["C3"] = 12000
        ws["C4"] = 11000

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果 - 合并单元格的值应该填充到所有单元格
        assert result is not None
        assert "合并单元格测试" in result
        assert "技术部" in result
        assert "员工1" in result
        print(f"✓ 合并单元格 Excel 处理成功")
        print(f"处理结果预览:\n{result[:400]}...")

    @pytest.mark.asyncio
    async def test_excel_chunking(self, tmp_path):
        """测试 Excel 每 10 行分块处理"""
        test_file = tmp_path / "test_chunking.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "分块测试"

        # 添加表头
        ws["A1"] = "编号"
        ws["B1"] = "数据"

        # 添加 25 行数据（应该分成 3 块：10+10+5）
        for i in range(1, 26):
            ws[f"A{i+1}"] = i
            ws[f"B{i+1}"] = f"数据{i}"

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果 - 应该包含多个数据行块
        assert result is not None
        assert "数据行 1-10" in result  # 第一块
        assert "数据行 11-20" in result  # 第二块
        assert "数据行 21-25" in result  # 第三块
        print(f"✓ Excel 分块处理成功（25行 -> 3块）")

    @pytest.mark.asyncio
    async def test_excel_duplicate_columns(self, tmp_path):
        """测试包含重复列名的 Excel 处理"""
        test_file = tmp_path / "test_duplicate_cols.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "重复列名"

        # 添加重复的表头
        ws["A1"] = "数据"
        ws["B1"] = "数据"
        ws["C1"] = "数据"

        # 添加数据
        ws["A2"] = "值1"
        ws["B2"] = "值2"
        ws["C2"] = "值3"

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果 - 重复列名应该被重命名
        assert result is not None
        # 检查是否包含原始数据
        assert "值1" in result or "值2" in result
        print(f"✓ 重复列名 Excel 处理成功")

    @pytest.mark.asyncio
    async def test_excel_multiple_sheets(self, tmp_path):
        """测试包含多个工作表的 Excel 处理"""
        test_file = tmp_path / "test_multi_sheets.xlsx"
        wb = Workbook()

        # 第一个工作表
        ws1 = wb.active
        ws1.title = "表1"
        ws1["A1"] = "列A"
        ws1["A2"] = "数据1"

        # 第二个工作表
        ws2 = wb.create_sheet("表2")
        ws2["A1"] = "列B"
        ws2["A2"] = "数据2"

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果 - 应该包含两个工作表
        assert result is not None
        assert "表1" in result
        assert "表2" in result
        assert "数据1" in result
        assert "数据2" in result
        print(f"✓ 多工作表 Excel 处理成功")

    @pytest.mark.asyncio
    async def test_excel_with_newlines(self, tmp_path):
        """测试包含换行符的单元格处理"""
        test_file = tmp_path / "test_newlines.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "换行测试"

        ws["A1"] = "标题"
        ws["A2"] = "第一行\n第二行\n第三行"

        wb.save(test_file)

        # 处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果 - 换行符应该被替换为 <br>
        assert result is not None
        assert "<br>" in result
        print(f"✓ 换行符处理成功（\\n -> <br>）")


class TestJSONProcessingChanges:
    """测试 JSON 处理的异步变更"""

    @pytest.mark.asyncio
    async def test_json_processing(self, tmp_path):
        """测试 JSON 文件的异步处理"""
        import json

        test_file = tmp_path / "test.json"
        test_data = {
            "name": "测试",
            "items": [{"id": 1, "value": "值1"}, {"id": 2, "value": "值2"}],
            "nested": {"key": "嵌套数据"},
        }

        # 写入 JSON 文件
        with open(test_file, "w", encoding="utf-8") as f:
            json.dump(test_data, f, ensure_ascii=False)

        # 异步处理文件
        result = await process_file_to_markdown(str(test_file))

        # 验证结果
        assert result is not None
        assert "```json" in result
        assert "测试" in result
        assert "嵌套数据" in result
        print(f"✓ JSON 异步处理成功")
        print(f"处理结果预览:\n{result[:200]}...")


if __name__ == "__main__":
    exit_code = pytest.main([__file__, "-v", "-s", "--tb=short", "--color=yes"])
    sys.exit(exit_code)
